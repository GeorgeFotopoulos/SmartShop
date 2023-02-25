import os

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


def create_excel(data):
    """
    Creates an excel and saves it on Desktop folder.

    Parameters:
        data (DataFrame): The data that will be saved in the Excel file.
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for row in dataframe_to_rows(data, index=False, header=True):
        worksheet.append(row)

    worksheet.auto_filter.ref = f"A1:{get_column_letter(data.shape[1])}1"
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 1
        worksheet.column_dimensions[column_letter].width = adjusted_width

    desktop_path = os.path.join(os.path.join(os.environ["USERPROFILE"]), "Desktop")
    workbook.save(desktop_path + "/products.xlsx")
